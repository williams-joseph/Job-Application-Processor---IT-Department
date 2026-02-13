"""
Excel export functionality for offline use.
"""

import os
from pathlib import Path
from typing import List, Dict, Optional
import logging

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    load_workbook = None
    Workbook = None

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Handles Excel file operations."""
    
    HEADER_ROW = [
        'S/N', 'NAME', 'POSITION CODE', 'GENDER', 'INT/EXT', 'DOB', 
        'AGE', 'NATIONALITY', 'EXP START (YEAR)', 'EXPERIENCE(Years)', 'QUALIFICATIONS'
    ]
    
    def __init__(self):
        if not load_workbook:
            raise ImportError("openpyxl not installed")
        from config import EXCEL_SHEET_NAME
        self.sheet_name = EXCEL_SHEET_NAME
    
    def append_to_excel(self, excel_path: str, data_rows: List[Dict]) -> Dict:
        """
        Update existing rows or append new ones in the Excel file.
        Inherits Position Code and Int/Ext from the existing data in the sheet.
        """
        excel_path = Path(excel_path)
        
        # Load or create workbook
        if excel_path.exists():
            wb = load_workbook(excel_path)
            if self.sheet_name in wb.sheetnames:
                ws = wb[self.sheet_name]
            else:
                ws = wb.active
            logger.info(f"Loaded existing Excel file: {excel_path}")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            self._write_header(ws)
            logger.info(f"Created new Excel file: {excel_path}")
        
        # Capture "Master" Position Code and Int/Ext from the first data row (Row 2)
        # This allows us to maintain consistency for this specific position/sheet
        master_pos_code = None
        master_int_ext = None
        
        if ws.max_row >= 2:
            # Column C (3) is POSITION CODE, Column E (5) is INT/EXT
            master_pos_code = ws.cell(row=2, column=3).value
            master_int_ext = ws.cell(row=2, column=5).value
            logger.info(f"Inheriting Position Code: '{master_pos_code}' and Status: '{master_int_ext}' from sheet.")

        # Create a mapping of Names to Row Numbers for fast lookup
        name_map = {}
        empty_rows = [] # Rows with position code but no name
        for row_idx in range(2, ws.max_row + 1):
            name_val = ws.cell(row=row_idx, column=2).value # NAME is Column B
            if name_val:
                name_map[str(name_val).strip().lower()] = row_idx
            else:
                pos_val = ws.cell(row=row_idx, column=3).value # POS CODE is Column C
                if pos_val:
                    empty_rows.append(row_idx)
        
        rows_updated = 0
        rows_appended = 0
        
        for data in data_rows:
            fields = data.get('fields', {})
            applicant_name = str(data.get('applicant_name', fields.get('NAME', ''))).strip().lower()
            
            if applicant_name in name_map:
                row_num = name_map[applicant_name]
                rows_updated += 1
            elif empty_rows:
                # Reuse the first available empty row that has a position code
                row_num = empty_rows.pop(0)
                # Set Name
                ws.cell(row=row_num, column=2).value = fields.get('NAME', '').upper()
                rows_updated += 1
            else:
                row_num = ws.max_row + 1
                ws.cell(row=row_num, column=1).value = row_num - 1 # S/N
                # Set Name if appending
                ws.cell(row=row_num, column=2).value = fields.get('NAME', '').upper()
                rows_appended += 1
            
            # 1. POSITION CODE (C) - Always use master value if found in sheet
            if master_pos_code:
                ws.cell(row=row_num, column=3).value = master_pos_code
            elif fields.get('POSITION CODE'):
                ws.cell(row=row_num, column=3).value = fields.get('POSITION CODE')
                
            # 2. GENDER (D)
            ws.cell(row=row_num, column=4).value = fields.get('GENDER', '')
            
            # 3. INT/EXT (E) - Always use master value if found in sheet
            if master_int_ext:
                ws.cell(row=row_num, column=5).value = master_int_ext
            elif fields.get('INT/EXT'):
                ws.cell(row=row_num, column=5).value = fields.get('INT/EXT')

            # 4. DOB (F)
            ws.cell(row=row_num, column=6).value = fields.get('DOB', '')
            # 5. AGE (G)
            ws.cell(row=row_num, column=7).value = fields.get('AGE', '')
            # 6. NATIONALITY (H)
            ws.cell(row=row_num, column=8).value = fields.get('NATIONALITY', '')
            # 7. EXP START (I)
            ws.cell(row=row_num, column=9).value = fields.get('EXP START (YEAR)', '')
            # 8. EXPERIENCE (J)
            ws.cell(row=row_num, column=10).value = fields.get('EXPERIENCE(Years)', '')
            
            # 9. QUALIFICATIONS (K)
            qual_cell = ws.cell(row=row_num, column=11)
            qual_cell.value = fields.get('QUALIFICATIONS', '')
            qual_cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Save workbook
        wb.save(excel_path)
        logger.info(f"Sync complete. Updated {rows_updated} rows and appended {rows_appended} rows")
        
        return {
            'rows_added': rows_updated + rows_appended,
            'rows_updated': rows_updated,
            'rows_appended': rows_appended,
            'file_path': str(excel_path),
        }
    
    def _write_header(self, worksheet):
        """Write header row with formatting."""
        for idx, header in enumerate(self.HEADER_ROW, start=1):
            cell = worksheet.cell(row=1, column=idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        col_widths = {
            'A': 8,   # S/N
            'B': 25,  # NAME
            'C': 20,  # POSITION CODE
            'D': 10,  # GENDER
            'E': 10,  # INT/EXT
            'F': 15,  # DOB
            'G': 8,   # AGE
            'H': 15,  # NATIONALITY
            'I': 18,  # EXP START
            'J': 18,  # EXPERIENCE
            'K': 50,  # QUALIFICATIONS
        }
        for col, width in col_widths.items():
            worksheet.column_dimensions[col].width = width
    
    def create_template(self, output_path: str):
        """Create a template Excel file with headers."""
        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name
        self._write_header(ws)
        wb.save(output_path)
        logger.info(f"Created template Excel file: {output_path}")

